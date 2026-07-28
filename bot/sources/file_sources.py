"""File-backed question sources: JSON, CSV and XLSX.

These are the importers an operator actually uses. Reading is delegated to worker
threads because the stdlib file, csv and openpyxl APIs are blocking, and a 1200
row spreadsheet parsed on the event loop would stall the bot's polling for the
duration.
"""

from __future__ import annotations

import asyncio
import csv
import json
from collections.abc import AsyncIterator, Iterator
from pathlib import Path
from typing import Any

from bot.sources.base import (
    QuestionSource,
    QuestionValidationError,
    RawQuestion,
    SourceError,
    parse_record,
)
from bot.utils.logging import get_logger

logger = get_logger(__name__)

#: Keys a JSON document may use to hold the list of questions.
_JSON_CONTAINER_KEYS = ("questions", "data", "items", "records", "tests")

#: Largest file we will read, as a guard against an accidental huge upload.
MAX_FILE_BYTES = 64 * 1024 * 1024


class FileQuestionSource(QuestionSource):
    """Common behaviour for sources backed by a single local file."""

    def __init__(
        self,
        path: Path,
        *,
        name: str | None = None,
        default_language: str = "uz",
        strict: bool = False,
    ) -> None:
        """Prepare a file source.

        Args:
            path: File to read.
            name: Source identity stored on every imported question. Defaults to
                ``"<format>:<filename stem>"``, which stays stable across
                re-imports of the same file.
            default_language: Language assumed when a record does not specify one.
            strict: Abort the whole import on the first bad record instead of
                skipping it. Off by default — one malformed row out of 1200
                should not cost the other 1199.
        """
        self.path = path
        self.default_language = default_language
        self.strict = strict
        self.name = name or f"{self.format_name}:{path.stem}"

        #: Records rejected during the last :meth:`fetch`, for the import report.
        self.errors: list[str] = []

    #: Short format label used in the default source name.
    format_name: str = "file"

    def _validate_readable(self) -> None:
        """Check the file exists and is a sane size before parsing.

        Raises:
            SourceError: The file is missing, empty or implausibly large.
        """
        if not self.path.exists():
            raise SourceError(f"File not found: {self.path}")
        if not self.path.is_file():
            raise SourceError(f"Not a file: {self.path}")
        size = self.path.stat().st_size
        if size == 0:
            raise SourceError(f"File is empty: {self.path.name}")
        if size > MAX_FILE_BYTES:
            raise SourceError(
                f"File is too large ({size / 1_048_576:.1f} MB, limit "
                f"{MAX_FILE_BYTES / 1_048_576:.0f} MB): {self.path.name}"
            )

    def _read_records(self) -> list[dict[str, Any]]:
        """Parse the file into raw mappings. Runs in a worker thread."""
        raise NotImplementedError

    async def count_estimate(self) -> int | None:
        """Number of records in the file, or ``None`` if it cannot be read."""
        try:
            self._validate_readable()
            records = await asyncio.to_thread(self._read_records)
        except SourceError:
            return None
        return len(records)

    async def fetch(self) -> AsyncIterator[RawQuestion]:
        """Yield every valid question in the file.

        Invalid records are collected in :attr:`errors` and skipped unless the
        source was constructed with ``strict=True``.

        Raises:
            SourceError: The file itself is unreadable.
            QuestionValidationError: A record was invalid and ``strict`` is set.
        """
        self._validate_readable()
        self.errors = []

        records = await asyncio.to_thread(self._read_records)
        logger.info("Read %d record(s) from %s", len(records), self.path.name)

        for position, record in enumerate(records, start=1):
            location = f"{self.path.name}:{position}"
            try:
                yield parse_record(
                    record,
                    location=location,
                    default_language=self.default_language,
                    # Position is a stable fallback id: the same row maps to the
                    # same question on re-import, keeping the import idempotent
                    # even for files with no id column.
                    fallback_id=str(position),
                )
            except QuestionValidationError as exc:
                if self.strict:
                    raise
                self.errors.append(str(exc))
                logger.warning("Skipping invalid record — %s", exc)


class JsonQuestionSource(FileQuestionSource):
    """Reads a ``.json`` file.

    Accepts either a bare array of questions or an object wrapping them under
    ``questions``/``data``/``items``/``records``/``tests``. An enclosing object
    may also carry ``source`` and ``language`` defaults.
    """

    format_name = "json"

    def _read_records(self) -> list[dict[str, Any]]:
        try:
            with self.path.open(encoding="utf-8") as handle:
                document = json.load(handle)
        except UnicodeDecodeError as exc:
            raise SourceError(f"{self.path.name} is not valid UTF-8 text") from exc
        except json.JSONDecodeError as exc:
            raise SourceError(
                f"{self.path.name} is not valid JSON (line {exc.lineno}, column {exc.colno}): "
                f"{exc.msg}"
            ) from exc

        if isinstance(document, dict):
            if (declared := document.get("language")) and isinstance(declared, str):
                self.default_language = declared.strip().lower()[:8]
            if (declared_name := document.get("source")) and isinstance(declared_name, str):
                self.name = f"json:{declared_name.strip()}"

            for key in _JSON_CONTAINER_KEYS:
                if isinstance(document.get(key), list):
                    document = document[key]
                    break
            else:
                raise SourceError(
                    f"{self.path.name}: expected a list of questions, or an object with "
                    f"one of {', '.join(_JSON_CONTAINER_KEYS)}"
                )

        if not isinstance(document, list):
            raise SourceError(f"{self.path.name}: top level must be a list or an object")

        records = [item for item in document if isinstance(item, dict)]
        if not records:
            raise SourceError(f"{self.path.name}: contains no question objects")
        return records


class CsvQuestionSource(FileQuestionSource):
    """Reads a ``.csv`` / ``.tsv`` file with a header row.

    The delimiter is sniffed, so comma-, semicolon- and tab-separated exports all
    work — Excel in a ru/uz locale writes semicolons, and demanding commas would
    reject the most likely file an operator has.
    """

    format_name = "csv"

    def _read_records(self) -> list[dict[str, Any]]:
        try:
            text = self.path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            # Windows-1251 is the common fallback for Cyrillic exports.
            try:
                text = self.path.read_text(encoding="cp1251")
                logger.warning("%s is not UTF-8; read it as cp1251", self.path.name)
            except UnicodeDecodeError as exc:
                raise SourceError(
                    f"{self.path.name}: cannot decode as UTF-8 or CP1251. "
                    "Re-save the file as UTF-8."
                ) from exc

        sample = text[:8192]
        try:
            dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(
                sample, delimiters=",;\t|"
            )
        except csv.Error:
            dialect = csv.excel
            logger.warning("%s: could not sniff delimiter, assuming comma", self.path.name)

        reader = csv.DictReader(text.splitlines(), dialect=dialect)
        if not reader.fieldnames:
            raise SourceError(f"{self.path.name}: missing a header row")

        records = [{key: value for key, value in row.items() if key is not None} for row in reader]
        if not records:
            raise SourceError(f"{self.path.name}: has a header but no data rows")
        return records


class XlsxQuestionSource(FileQuestionSource):
    """Reads the first worksheet of an ``.xlsx`` workbook.

    The first row is treated as the header. Opened read-only so a large workbook
    streams instead of being materialised in full.
    """

    format_name = "xlsx"

    def _read_records(self) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is declared
            raise SourceError("openpyxl is required to read .xlsx files") from exc

        try:
            workbook = load_workbook(self.path, read_only=True, data_only=True)
        except Exception as exc:
            raise SourceError(f"{self.path.name}: cannot open as a workbook ({exc})") from exc

        try:
            worksheet = workbook.worksheets[0]
            rows: Iterator[tuple[Any, ...]] = worksheet.iter_rows(values_only=True)

            try:
                header_row = next(rows)
            except StopIteration as exc:
                raise SourceError(f"{self.path.name}: worksheet is empty") from exc

            headers = [
                str(cell).strip() if cell is not None else f"column_{index}"
                for index, cell in enumerate(header_row, start=1)
            ]

            records: list[dict[str, Any]] = []
            for row in rows:
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue  # blank spacer row
                records.append(
                    {headers[index]: cell for index, cell in enumerate(row) if index < len(headers)}
                )
        finally:
            workbook.close()

        if not records:
            raise SourceError(f"{self.path.name}: has a header but no data rows")
        return records
